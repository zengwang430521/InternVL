import os
import re
import json
import argparse
import xlrd
from openpyxl import Workbook, load_workbook
from tqdm import tqdm


def process_line(line):
    convs = []

    if 'INFO [IMAGE]:' in line:
        text = line.split('INFO [IMAGE]:')[1]
        content = eval(text)
        convs.append({'role':'image', 'content': content})

    if 'INFO [PROMPT]:' in line:
        text = line.split('INFO [PROMPT]:')[1]
        text = eval(text)
        pattern = r'<\|im_start\|>(.*?)<\|im_end\|>'
        matches = re.findall(pattern, text, re.DOTALL)
        if len(matches):
            for match in matches:
                role = match.split('\n')[0]
                content = match.split(role)[1].strip()
                convs.append({'role':role, 'content':content})

    if "INFO [RESPONSE]:" in line:
        text = line.split('INFO [RESPONSE]:')[1]
        tmp_dict = eval(text)
        convs.append({'role': 'real_assistant', 'content': tmp_dict['answer']})

    return convs


log_file = '/home/SENSETIME/zengwang/myprojects/llm/agent/test_data/SenseChat-Vision-0903/model_worker_worker_180398cc.log'
with open(log_file, 'r') as f:
    lines = f.readlines()
convs = []
for line in tqdm(lines):
    convs += process_line(line)

output_file = '/home/SENSETIME/zengwang/myprojects/llm/agent/test_data/SenseChat-Vision-0903/model_worker_worker_180398cc.xlsx'


'保存'
wb = Workbook()
default_sheet = wb.active
wb.remove(default_sheet)
sheet1 = wb.create_sheet('结果')
sheet1['A1'] = 'image'
sheet1['B1'] = 'system'
sheet1['C1'] = 'user'
sheet1['D1'] = 'response'
# sheet1['E2'] = 'model result'

image = ''
system_prompt = ''
user_prompt = ''
response = ''
idx_row = 2

for conv in tqdm(convs):
    role, content = conv['role'], conv['content']
    if role == 'image':
        image = content[0] if len(content) else ''
    elif role == 'system':
        system_prompt = content
    elif role == 'user':
        user_prompt = f'{user_prompt}{content}'
    elif role == 'real_assistant':
        response = content

        sheet1[f'A{idx_row}'] = image
        sheet1[f'B{idx_row}'] = system_prompt
        sheet1[f'C{idx_row}'] = user_prompt
        sheet1[f'D{idx_row}'] = response

        image = ''
        system_prompt = ''
        user_prompt = ''
        response = ''
        idx_row += 1
    else:
        print(role)
wb.save(output_file)





t = 0